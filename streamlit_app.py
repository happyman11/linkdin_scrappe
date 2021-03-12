
    
from linkedin_scraper import Person, actions
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import streamlit as st
import streamlit.components.v1 as components
import numpy as np
import pandas as pd
import openpyxl 
import base64

def filedownload(df):
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()  # strings <-> bytes conversions
    href = f'<a href="data:file/csv;base64,{b64}" download="Profile_data.csv">Download CSV File</a>'
    return href


def read_pdf(path):
    
    if path is not None:
     
        List_email = openpyxl.load_workbook(path)
        
        return(List_email)

def get_information_bulk(data):
    
    chrome_options = Options()  
    chrome_options.add_argument("--headless")
    
    path_webdriver="/home/ravishekhartiwari/Desktop/LINKDIN_SCRAPP/chromedriver_linux64/chromedriver"
    driver = webdriver.Chrome(executable_path=path_webdriver,chrome_options=chrome_options)


    
    #email = ""
    #password = ""
    #actions.login(driver,email,password) # if email and password isnt given, it'll prompt in terminal
    actions.login(driver)
    
    sheet_obj = data.active 
    m_row = sheet_obj.max_row 
    count=0
    
    name=[]
    experiences=[]
    education=[]
    company=[]
    job_title=[]
    about=[]
    linkdin_url=[]
    

    
    for i in range(1, m_row + 1): 
        cell_obj = sheet_obj.cell(row = i, column = 1)
        url=cell_obj.value
        
        person = Person((str(url)).strip(), driver=driver) 
        
        name.append(person.name)
        experiences.append(person.experiences)
        education.append(person.educations)
        company.append(person.company)
        job_title.append(person.job_title)
        about.append(person.about)
        linkdin_url.append(person.linkedin_url)

    #driver.close()

    return(name,experiences,education,company,job_title,about,linkdin_url,driver)

def for_single_url(url):
    
    
    chrome_options = Options()  
    chrome_options.add_argument("--headless")
    
    path_webdriver="/home/ravishekhartiwari/Desktop/LINKDIN_SCRAPP/chromedriver_linux64/chromedriver"
    driver = webdriver.Chrome(executable_path=path_webdriver,chrome_options=chrome_options)


    #email = ""
    #password = ""
    #actions.login(driver,email,password) # if email and password isnt given, it'll prompt in terminal
    actions.login(driver)

    person = Person(url, driver=driver)    

    person_details={ "name" :person.name,
                     "experiences": person.experiences,
                     "education" :person.educations,
                     "company" : person.company,
                     "job_title":person.job_title,
                     "about" :person.about,
                     "linkdin_url" :person.linkedin_url}
    



    #person.scrape(close_on_complete=True)
    #driver.close()
    
    return (person_details)    

st.title("***User Information Extraction on Linkdin***") 

st.sidebar.header("User Input")

st.sidebar.subheader("Feed url of the Person")

url_person = st.sidebar.text_input('Linkdin Url', 'Url')

if(st.sidebar.button("Get information")):
    
    sanitised_url=str(url_person).strip()
    person_details=for_single_url(sanitised_url)
    
    
    details=pd.DataFrame({
            "Name" :person_details["name"] ,
            "Education" : person_details["education"],
            "Company" :   person_details["company"],
             "Job_title" : person_details["job_title"],
             "About"  :person_details["about"],
             "Linkdin_url"  :person_details["linkdin_url"],
            
            })
    
    
    st.write(details)

    st.markdown(filedownload(details), unsafe_allow_html=True)

st.sidebar.subheader("Linkdin Profile(.xlsx format only)")
uploaded_file = st.sidebar.file_uploader("Upload your input file", type=["xlsx"])

if(st.sidebar.button("GEt information of Uploaded file")):

    emails_list=read_pdf(uploaded_file)
    name,experiences,education,company,job_title,about,linkdin_url,driver=get_information_bulk(emails_list)
   

    details=pd.DataFrame({
            "Name" :name ,
            "Education" : education,
            "Company" :   company,
            "Job_title" :job_title,
            "About"  :about,
            "Linkdin_url"  :linkdin_url,
            
            })

    st.write(details)

    st.markdown(filedownload(details), unsafe_allow_html=True)



